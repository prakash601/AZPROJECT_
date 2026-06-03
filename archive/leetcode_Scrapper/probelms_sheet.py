# import pandas as pd

# # Read the data from Qindex.txt and index.txt
# with open('leetcode_Scrapper/Qindex.txt', 'r') as q_file:
#     q_data = q_file.read().splitlines()

# with open('leetcode_Scrapper/index.txt', 'r') as index_file:
#     index_data = index_file.read().splitlines()

# # Create a DataFrame with the data
# data = {'Question Number': q_data, 'Question Name': index_data}
# df = pd.DataFrame(data)

# # Write the DataFrame to an Excel file
# df.to_excel('questions.xlsx', index=False)
import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook('questions.xlsx')

# Select the specific worksheet where you want to delete rows
worksheet = workbook['Sheet2']  # Replace 'Sheet1' with the actual sheet name

# Create a list of numbers you want to keep
numbers_to_keep = [1832, 38, 692, 12, 219, 76, 645, 1239, 1662, 523, 835, 49, 2136, 1293, 766, 1706, 433, 2131, 354, 212, 899, 1323, 1544, 901, 1047, 26, 295, 151, 947, 222, 374, 223, 263, 587, 224, 1926, 279, 36, 79, 907, 1235, 446, 2225, 380, 1207, 1704, 1657, 451, 2256, 876, 328, 938, 872, 1026, 1339, 124, 70, 931, 198, 1143, 232, 150, 739, 1971, 841, 886, 834, 309, 790, 2389, 55, 2279, 1962, 1834, 797, 980, 290, 520, 944, 2244, 452, 1833, 134, 149, 144, 100, 1443, 1519, 2246, 1061, 2421, 57, 926, 918, 974, 491, 93, 131, 997, 909, 2359, 787, 472, 352, 460, 1137, 1626, 1071, 953, 6, 567, 438, 1470, 904, 45, 2306, 1162, 1129, 2477, 1523, 67, 989, 104, 783, 226, 103, 35, 540, 1011, 502, 1675, 121, 72, 427, 652, 912, 443, 28, 2444, 1345, 1539, 2187, 875, 142, 382, 109, 23, 101, 129, 958, 106, 208, 1472, 211, 605, 2348, 2492, 1319, 1466, 2316, 2360, 64, 983, 1402, 87, 1444, 704, 2300, 881, 2405, 2439, 1254, 1020, 133, 1857, 20, 2390, 71, 946, 576, 2218, 1639, 1431, 1768, 1372, 662, 879, 1312, 1416, 1046, 2336, 258, 319, 839, 1697, 1579, 1491, 1822, 2215, 649, 1456, 1498, 1964, 1572, 54, 59, 1035, 2140, 2466, 1799, 1721, 24, 2130, 1557, 785, 399, 934, 347, 703, 2542, 837, 1140, 1406, 1547, 1603, 705, 1396, 1091, 2101, 1376, 547, 1232, 1502, 1318, 1351, 744, 1802, 1146, 228, 2352, 530, 1161, 1569, 1187, 2328, 1732, 2090, 2448, 714, 1027, 956, 1575, 2462, 373, 1514, 864, 1970, 2305, 1601, 859, 137, 1493, 209, 2024, 2551, 2272, 111, 863, 802, 207, 1218, 1751, 1125, 445, 146, 435, 735, 673, 688, 894, 50, 852, 1870, 2141, 486, 808, 664, 712, 77, 46, 67, 139, 95, 920, 74, 33, 2616, 81, 518, 63, 2369, 215, 86, 239, 542, 1615, 1489, 1203, 459, 168, 767, 68, 97, 646, 403, 225, 2483, 2366, 1326, 338, 2707, 62, 141, 138, 725, 92, 118, 377, 1359, 1282, 1647, 135, 332, 1584, 1631, 847, 1337, 287, 1658, 4]

# Convert the generator to a list and iterate through the rows
rows_to_delete = []
for row in list(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1)):  # Assuming Column A
    cell_value = row[0].value
    if cell_value not in numbers_to_keep:
        rows_to_delete.append(row[0].row)

# Delete the rows
for row_number in reversed(rows_to_delete):
    worksheet.delete_rows(row_number)

# Save the modified workbook
workbook.save('modified_excel_file.xlsx')

# Close the workbook
workbook.close()
